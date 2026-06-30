from __future__ import annotations

import math
import re
import zipfile
from collections import defaultdict
from collections import Counter
from copy import copy
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

try:
    import xlrd
except ImportError:  # pragma: no cover - only hit when dependency install is broken
    xlrd = None


NUMBER_RE = re.compile(r"-?\d+(?:\.\d+)?")
CELL_REF_RE = re.compile(r"\$?([A-Z]{1,3})\$?(\d+)")


@dataclass
class TableMap:
    header_row: int
    data_start: int
    columns: dict[str, int]


class SpreadsheetReadError(ValueError):
    pass


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value)).strip()


def normalize_key(value: Any) -> str:
    text = normalize_text(value).lower()
    return re.sub(r"[^\w\u4e00-\u9fff]+", "", text)


def clean_finished_name(value: Any) -> str:
    text = normalize_text(value)
    text = re.sub(r"投料单$", "", text)
    text = re.sub(r"（停用）|\(停用\)", "", text)
    return text.strip()


def to_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if text.startswith("="):
        return None
    match = NUMBER_RE.search(text.replace(",", ""))
    return float(match.group()) if match else None


def display_number(value: float | None) -> float | int | None:
    if value is None:
        return None
    return int(value) if float(value).is_integer() else round(float(value), 4)


def first_present(*values: Any) -> Any:
    for value in values:
        if value not in (None, ""):
            return value
    return ""


def last_nonempty_row(ws) -> int:
    rows = [row for (row, _col), cell in ws._cells.items() if cell.value not in (None, "")]
    return max(rows, default=1)


def _header_text(ws, row: int, col: int) -> str:
    parts = []
    for r in (row, row + 1):
        if r <= ws.max_row:
            value = ws.cell(r, col).value
            if value not in (None, ""):
                parts.append(str(value))
    return normalize_text("".join(parts))


def _find_col(headers: dict[int, str], keywords: Iterable[str]) -> int | None:
    for keyword in keywords:
        for col, text in headers.items():
            if keyword in text:
                return col
    return None


def _table_columns_from_headers(headers: dict[int, str], current_headers: dict[int, str], purpose: str) -> dict[str, int] | None:
    product = _find_col(current_headers, ["商品名称", "原料名称", "存货名称", "产品名称"])
    if not product:
        return None

    columns: dict[str, int] = {"product": product}
    pairs = {
        "sequence": ["序号"],
        "category": ["类别", "备注"],
        "code": ["存货编码", "产品编码", "编码"],
        "spec": ["规格型号", "商品规格", "规格"],
        "unit": ["主计量", "单位"],
        "price": ["本币无税单价", "无税单价", "含税单价", "单价"],
        "order_qty": ["订货数量"],
        "qty": ["数量"],
        "safety": ["安全库存数", "安全库存"],
        "inventory": ["盘点库存数", "盘点库存"],
        "inbound": ["入库数", "入库"],
        "outbound": ["出库数量", "出库数", "订货数量"],
        "theory_stock": ["理论库存数", "理论库存"],
        "production": ["理论排产", "排产量", "产量"],
        "warehouse": ["所属库", "库位", "仓库", "库别"],
    }
    for key, keywords in pairs.items():
        col = _find_col(headers, keywords)
        if col:
            columns[key] = col

    if purpose == "order" and "order_qty" not in columns and "qty" in columns:
        columns["order_qty"] = columns["qty"]
    if purpose == "material" and "qty" not in columns and "order_qty" in columns:
        columns["qty"] = columns["order_qty"]
    return columns


def detect_table(ws, purpose: str = "generic") -> TableMap | None:
    scan_rows = min(last_nonempty_row(ws), 30)
    for row in range(1, scan_rows + 1):
        current_headers = {col: normalize_text(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)}
        headers = {col: _header_text(ws, row, col) for col in range(1, ws.max_column + 1)}
        columns = _table_columns_from_headers(headers, current_headers, purpose)
        if not columns:
            continue

        return TableMap(header_row=row, data_start=row + 1, columns=columns)
    return None


def _is_legacy_xls(path: Path) -> bool:
    return path.suffix.lower() == ".xls"


def _format_xls_value(value: Any) -> Any:
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _open_xls_workbook(path: Path):
    if xlrd is None:
        raise SpreadsheetReadError(f"{path.name} 是老 .xls 格式，当前服务缺少 xlrd 依赖，无法读取。")
    try:
        return xlrd.open_workbook(str(path))
    except Exception as exc:
        raise SpreadsheetReadError(f"{path.name} 读取失败：{exc}") from exc


def _last_nonempty_xls(sheet) -> int:
    for row_idx in range(sheet.nrows - 1, -1, -1):
        if any(normalize_text(_format_xls_value(sheet.cell_value(row_idx, col_idx))) for col_idx in range(sheet.ncols)):
            return row_idx + 1
    return 0


def _header_text_xls(sheet, row: int, col: int) -> str:
    parts = []
    for row_idx in (row - 1, row):
        if 0 <= row_idx < sheet.nrows and 0 <= col - 1 < sheet.ncols:
            value = _format_xls_value(sheet.cell_value(row_idx, col - 1))
            if value not in (None, ""):
                parts.append(str(value))
    return normalize_text("".join(parts))


def _detect_table_xls(sheet, purpose: str = "generic") -> TableMap | None:
    scan_rows = min(_last_nonempty_xls(sheet), 30)
    for row in range(1, scan_rows + 1):
        current_headers = {
            col: normalize_text(_format_xls_value(sheet.cell_value(row - 1, col - 1)))
            for col in range(1, sheet.ncols + 1)
        }
        headers = {col: _header_text_xls(sheet, row, col) for col in range(1, sheet.ncols + 1)}
        columns = _table_columns_from_headers(headers, current_headers, purpose)
        if columns:
            return TableMap(header_row=row, data_start=row + 1, columns=columns)
    return None


def infer_store_name(ws, path: Path) -> str:
    for cell in ("A1", "B1", "A2", "B2"):
        value = ws[cell].value
        if not value:
            continue
        text = str(value)
        match = re.search(r"馄饨侯[（(]?([^）)店]+)[）)]?店", text)
        if match:
            return f"{match.group(1)}店"
    title = ws.title.replace("订货", "").strip()
    if title and not title.lower().startswith("sheet"):
        return title if title.endswith(("店", "学校")) else f"{title}店"
    return path.stem


def infer_store_name_xls(sheet, path: Path) -> str:
    for row_idx, col_idx in ((0, 0), (0, 1), (1, 0), (1, 1)):
        if row_idx >= sheet.nrows or col_idx >= sheet.ncols:
            continue
        value = _format_xls_value(sheet.cell_value(row_idx, col_idx))
        if not value:
            continue
        text = str(value)
        match = re.search(r"馄饨侯[（(]?([^）)店]+)[）)]?店", text)
        if match:
            return f"{match.group(1)}店"
    title = str(sheet.name).replace("订货", "").strip()
    if title and not title.lower().startswith("sheet"):
        return title if title.endswith(("店", "学校")) else f"{title}店"
    return path.stem


def parse_rows_xls(path: Path, purpose: str = "generic") -> list[dict[str, Any]]:
    wb = _open_xls_workbook(path)
    parsed: list[dict[str, Any]] = []
    for sheet in wb.sheets():
        table = _detect_table_xls(sheet, purpose)
        if not table:
            continue
        store = infer_store_name_xls(sheet, path)
        columns = table.columns
        product_col = columns["product"]
        for row in range(table.data_start, _last_nonempty_xls(sheet) + 1):
            product = _format_xls_value(sheet.cell_value(row - 1, product_col - 1))
            if not normalize_text(product):
                continue
            item: dict[str, Any] = {
                "product": str(product).strip(),
                "product_key": normalize_key(product),
                "store": store,
                "source": path.name,
                "row": row,
            }
            for key in ("sequence", "category", "code", "spec", "unit", "price", "order_qty", "qty", "safety", "inventory", "inbound", "outbound", "theory_stock", "production", "warehouse"):
                col = columns.get(key)
                if not col:
                    continue
                value = _format_xls_value(sheet.cell_value(row - 1, col - 1))
                if key in {"price", "order_qty", "qty", "safety", "inventory", "inbound", "outbound", "theory_stock", "production"}:
                    item[key] = to_number(value)
                else:
                    item[key] = "" if value is None else str(value).strip()
            parsed.append(item)
    return parsed


def parse_rows(path: Path, purpose: str = "generic") -> list[dict[str, Any]]:
    if _is_legacy_xls(path):
        return parse_rows_xls(path, purpose)
    try:
        wb = load_workbook(path, data_only=True)
    except InvalidFileException as exc:
        raise SpreadsheetReadError(f"{path.name} 格式暂不支持。请把文件另存为 .xlsx 后重新上传。") from exc
    parsed: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        table = detect_table(ws, purpose)
        if not table:
            continue
        store = infer_store_name(ws, path)
        columns = table.columns
        for row in range(table.data_start, last_nonempty_row(ws) + 1):
            product = ws.cell(row, columns["product"]).value
            if not normalize_text(product):
                continue
            item: dict[str, Any] = {
                "product": str(product).strip(),
                "product_key": normalize_key(product),
                "store": store,
                "source": path.name,
                "row": row,
            }
            for key in ("sequence", "category", "code", "spec", "unit", "price", "order_qty", "qty", "safety", "inventory", "inbound", "outbound", "theory_stock", "production", "warehouse"):
                col = columns.get(key)
                if not col:
                    continue
                value = ws.cell(row, col).value
                if key in {"price", "order_qty", "qty", "safety", "inventory", "inbound", "outbound", "theory_stock", "production"}:
                    item[key] = to_number(value)
                else:
                    item[key] = "" if value is None else str(value).strip()
            parsed.append(item)
    return parsed


def aggregate_orders(paths: list[Path], confirmed_items: list[dict[str, Any]] | None = None) -> dict[str, dict[str, Any]]:
    summary: dict[str, dict[str, Any]] = {}
    for path in paths:
        for row in parse_rows(path, "order"):
            qty = row.get("order_qty")
            if qty is None or qty == 0:
                continue
            key = row["product_key"]
            current = summary.setdefault(
                key,
                {
                    "product": row["product"],
                    "category": row.get("category", ""),
                    "code": row.get("code", ""),
                    "spec": row.get("spec", ""),
                    "unit": row.get("unit", ""),
                    "price": row.get("price"),
                    "quantity": 0.0,
                    "stores": defaultdict(float),
                },
            )
            current["quantity"] += float(qty)
            current["stores"][row["store"]] += float(qty)
            if current.get("price") is None and row.get("price") is not None:
                current["price"] = row.get("price")

    for item in confirmed_items or []:
        product = str(item.get("product") or item.get("name") or "").strip()
        if not product:
            continue
        qty = to_number(item.get("quantity"))
        if qty is None or qty == 0:
            continue
        key = normalize_key(product)
        price = to_number(item.get("price"))
        current = summary.setdefault(
            key,
            {
                "product": product,
                "category": item.get("category", ""),
                "code": item.get("code", ""),
                "spec": item.get("spec", ""),
                "unit": item.get("unit", ""),
                "price": price,
                "quantity": 0.0,
                "stores": defaultdict(float),
            },
        )
        current["quantity"] += float(qty)
        if current.get("price") is None and price is not None:
            current["price"] = price
        store = str(item.get("store") or "文字加单").strip()
        current["stores"][store] += float(qty)
    return summary


def _store_safety_value(result: dict[str, float], name: Any, qty: Any, code: Any = None) -> None:
    number = to_number(qty)
    if number is None:
        return
    product_key = normalize_key(name)
    if product_key:
        result[product_key] = float(number)
    code_key = normalize_key(code)
    if code_key:
        result[code_key] = float(number)


def _fallback_safety_rows(ws, result: dict[str, float]) -> None:
    last = last_nonempty_row(ws)
    for row in range(1, last + 1):
        cells = [(col, ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        text_cells = [
            (col, value)
            for col, value in cells
            if normalize_text(value) and (to_number(value) is None or re.search(r"[A-Za-z\u4e00-\u9fff]", str(value)))
        ]
        numeric_cells = [(col, to_number(value)) for col, value in cells if to_number(value) is not None]
        if not text_cells or not numeric_cells:
            continue
        name_col, name = max(text_cells, key=lambda item: len(normalize_text(item[1])))
        right_numbers = [(col, value) for col, value in numeric_cells if col > name_col]
        qty = (right_numbers or numeric_cells)[-1][1]
        _store_safety_value(result, name, qty)


def safety_stock_map(path: Path | None) -> dict[str, float]:
    if not path:
        return {}
    result: dict[str, float] = {}
    wb = load_workbook(path, data_only=True)
    for ws in wb.worksheets:
        table = detect_table(ws, "safety")
        if table and ("safety" in table.columns or "qty" in table.columns):
            product_col = table.columns["product"]
            safety_col = table.columns.get("safety") or table.columns.get("qty")
            code_col = table.columns.get("code")
            for row in range(table.data_start, last_nonempty_row(ws) + 1):
                _store_safety_value(
                    result,
                    ws.cell(row, product_col).value,
                    ws.cell(row, safety_col).value,
                    ws.cell(row, code_col).value if code_col else None,
                )
            continue

        header_row = None
        product_col = None
        safety_col = None
        code_col = None
        for row in range(1, min(last_nonempty_row(ws), 30) + 1):
            headers = {col: normalize_text(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)}
            product_col = _find_col(headers, ["商品名称", "产品名称", "物品名称", "存货名称", "原料名称", "品名", "名称"])
            safety_col = _find_col(headers, ["安全库存数", "安全库存", "库存标准", "标准库存", "安全量"])
            code_col = _find_col(headers, ["存货编码", "产品编码", "商品编码", "编码"])
            if product_col and safety_col:
                header_row = row
                break
        if header_row and product_col and safety_col:
            for row in range(header_row + 1, last_nonempty_row(ws) + 1):
                _store_safety_value(
                    result,
                    ws.cell(row, product_col).value,
                    ws.cell(row, safety_col).value,
                    ws.cell(row, code_col).value if code_col else None,
                )
            continue

        _fallback_safety_rows(ws, result)
    return result


def product_catalog(paths: list[Path]) -> dict[str, dict[str, Any]]:
    catalog: dict[str, dict[str, Any]] = {}
    for path in paths:
        if not path:
            continue
        for row in parse_rows(path, "generic"):
            key = row["product_key"]
            if key in catalog:
                continue
            catalog[key] = {
                "product": row["product"],
                "category": row.get("category", ""),
                "code": row.get("code", ""),
                "spec": row.get("spec", ""),
                "unit": row.get("unit", ""),
                "price": row.get("price"),
            }
    return catalog


def extract_receipt_template_skus(path: Path, limit: int = 1000) -> dict[str, Any]:
    seen: set[tuple[str, str, str, str]] = set()
    products: list[dict[str, str]] = []
    source_rows = 0
    unique_rows = 0
    for row in parse_rows(path, "receipt"):
        name = str(row.get("product") or "").strip()
        if not name:
            continue
        source_rows += 1
        item = {
            "name": name,
            "spec": str(row.get("spec") or "").strip(),
            "unit": str(row.get("unit") or "").strip(),
            "category": str(row.get("category") or "").strip(),
        }
        key = tuple(normalize_text(item[field]) for field in ("name", "spec", "unit", "category"))
        if key in seen:
            continue
        seen.add(key)
        unique_rows += 1
        if len(products) < limit:
            products.append(item)
    return {
        "products": products,
        "source_rows": source_rows,
        "unique_rows": unique_rows,
        "deduped": max(source_rows - unique_rows, 0),
        "truncated": max(unique_rows - len(products), 0),
        "limit": limit,
    }


def copy_row_format(ws, src_row: int, dst_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.border:
            dst.border = copy(src.border)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.font:
            dst.font = copy(src.font)


def clear_rows_without_shift(ws, start_row: int, end_row: int) -> None:
    if end_row < start_row:
        return
    for merged in list(ws.merged_cells.ranges):
        if start_row <= merged.min_row and merged.max_row <= end_row:
            ws.unmerge_cells(str(merged))
    for key in list(ws._cells):
        row, _col = key
        if start_row <= row <= end_row:
            del ws._cells[key]
    for row in range(start_row, end_row + 1):
        ws.row_dimensions.pop(row, None)


def trim_blank_tail(ws, keep_last_row: int) -> None:
    for key, cell in list(ws._cells.items()):
        row, _col = key
        if row > keep_last_row and cell.value in (None, ""):
            del ws._cells[key]
    for row in [row for row in ws.row_dimensions if row > keep_last_row]:
        del ws.row_dimensions[row]


def write_basic_headers(ws, title: str, headers: list[str]) -> None:
    ws.title = "Sheet1"
    ws.append([title])
    ws.append(headers)
    ws.freeze_panes = "A3"
    for idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(idx)].width = max(12, min(24, len(header) * 2))


def generate_production_workbook(
    order_paths: list[Path],
    production_template_path: Path | None,
    safety_stock_path: Path | None,
    confirmed_items: list[dict[str, Any]] | None,
    order_date: date | None,
    output_dir: Path,
) -> tuple[Path, list[str]]:
    output_dir.mkdir(parents=True, exist_ok=True)
    warnings: list[str] = []
    orders = aggregate_orders(order_paths, confirmed_items)
    safety_values = safety_stock_map(safety_stock_path)

    catalog_paths = [p for p in [production_template_path, *order_paths] if p]
    catalog = product_catalog(catalog_paths)
    for key, order in orders.items():
        catalog.setdefault(
            key,
            {
                "product": order["product"],
                "category": order.get("category", ""),
                "code": order.get("code", ""),
                "spec": order.get("spec", ""),
                "unit": order.get("unit", ""),
                "price": order.get("price"),
            },
        )

    keys = [key for key, order in orders.items() if order.get("quantity")]
    rows = []
    for idx, key in enumerate(keys, 1):
        order = orders.get(key, {})
        meta = catalog.get(key, order)
        order_qty = order.get("quantity", 0.0)
        code_key = normalize_key(meta.get("code", ""))
        safety = safety_values.get(key)
        if safety is None and code_key:
            safety = safety_values.get(code_key)
        rows.append(
            {
                "sequence": idx,
                "category": first_present(meta.get("category"), order.get("category")),
                "code": first_present(meta.get("code"), order.get("code")),
                "product": first_present(meta.get("product"), order.get("product")),
                "spec": first_present(meta.get("spec"), order.get("spec")),
                "unit": first_present(meta.get("unit"), order.get("unit")),
                "price": first_present(order.get("price"), meta.get("price")) or None,
                "inventory": None,
                "safety": safety,
                "inbound": None,
                "outbound": order_qty,
                "theory_stock_formula": None,
                "production": safety,
            }
        )

    workbook_date = order_date or date.today()
    if production_template_path:
        wb = load_workbook(production_template_path)
        ws = wb.worksheets[0]
        table = detect_table(ws, "production") or TableMap(
            header_row=2,
            data_start=4,
            columns={
                "sequence": 1,
                "category": 2,
                "product": 3,
                "spec": 4,
                "unit": 5,
                "price": 6,
                "inventory": 7,
                "safety": 8,
                "inbound": 9,
                "outbound": 10,
                "theory_stock": 11,
                "production": 12,
            },
        )
        max_col = max(12, ws.max_column)
        data_start = max(4, table.data_start)
        original_last = last_nonempty_row(ws)
        for r in range(data_start, data_start + len(rows)):
            copy_row_format(ws, data_start, r, max_col)
            for c in range(1, max_col + 1):
                ws.cell(r, c).value = None
        delete_start = data_start + len(rows)
        if original_last >= delete_start:
            clear_rows_without_shift(ws, delete_start, original_last)
        trim_blank_tail(ws, max(data_start + len(rows) - 1, table.header_row))
        cols = table.columns
        for cell in ("G2", "I2", "J2", "K2", "L2"):
            if ws[cell].value is not None:
                ws[cell].value = workbook_date
        for offset, item in enumerate(rows):
            r = data_start + offset
            values = {
                "sequence": item["sequence"],
                "category": item["category"],
                "code": item["code"],
                "product": item["product"],
                "spec": item["spec"],
                "unit": item["unit"],
                "price": display_number(item["price"]),
                "inventory": None,
                "safety": display_number(item["safety"]),
                "inbound": None,
                "outbound": display_number(item["outbound"]),
                "theory_stock": None,
                "production": display_number(item["production"]),
            }
            for key, value in values.items():
                col = cols.get(key)
                if col:
                    ws.cell(r, col).value = value
    else:
        wb = Workbook()
        ws = wb.active
        headers = ["序号", "类别", "商品名称", "商品规格", "单位", "单价（元）", "盘点库存数", "安全库存数", "入库数", "出库数量", "理论库存数", "排产量"]
        write_basic_headers(ws, "排产单", headers)
        for row_index, item in enumerate(rows, start=3):
            ws.append(
                [
                    item["sequence"],
                    item["category"],
                    item["product"],
                    item["spec"],
                    item["unit"],
                    display_number(item["price"]),
                    None,
                    display_number(item["safety"]),
                    None,
                    display_number(item["outbound"]),
                    None,
                    display_number(item["production"]),
                ]
            )
    output = output_dir / f"排产表_待补充_{workbook_date.isoformat()}.xlsx"
    wb.save(output)
    return output, warnings


def generate_completed_production_workbook(
    production_path: Path,
    document_date: date | None,
    output_dir: Path,
) -> tuple[Path, list[str]]:
    output_dir.mkdir(parents=True, exist_ok=True)
    warnings: list[str] = []
    wb = load_workbook(production_path, data_only=False)
    updated_rows = 0
    skipped_rows = 0
    for ws in wb.worksheets:
        table = detect_table(ws, "production")
        if not table:
            continue
        cols = table.columns
        theory_col = cols.get("theory_stock")
        production_col = cols.get("production")
        safety_col = cols.get("safety")
        inventory_col = cols.get("inventory")
        inbound_col = cols.get("inbound")
        outbound_col = cols.get("outbound")
        product_col = cols.get("product")
        if not all([theory_col, inventory_col, inbound_col, product_col]):
            warnings.append(f"{ws.title} 缺少盘点库存数、入库数或理论库存数列，已跳过。")
            continue
        for row in range(table.data_start, last_nonempty_row(ws) + 1):
            if not normalize_text(ws.cell(row, product_col).value):
                continue
            if production_col and safety_col:
                safety = to_number(ws.cell(row, safety_col).value)
                if safety is not None:
                    ws.cell(row, production_col).value = display_number(safety)
            inventory = to_number(ws.cell(row, inventory_col).value)
            inbound = to_number(ws.cell(row, inbound_col).value)
            outbound = to_number(ws.cell(row, outbound_col).value) if outbound_col else 0.0
            if inventory is None or inbound is None:
                skipped_rows += 1
                ws.cell(row, theory_col).value = None
                continue
            theory_stock = float(inventory) + float(inbound) - float(outbound or 0.0)
            ws.cell(row, theory_col).value = display_number(theory_stock)
            updated_rows += 1
    if not updated_rows:
        warnings.append("没有计算到理论库存数，请确认已上传填好盘点库存数和入库数的排产表。")
    if skipped_rows:
        warnings.append(f"有 {skipped_rows} 行缺少盘点库存数或入库数，理论库存数已留空。")
    output_date = (document_date or date.today()).isoformat()
    output = output_dir / f"排产表_{output_date}.xlsx"
    wb.save(output)
    return output, warnings


def _best_order_sheet(path: Path):
    wb = load_workbook(path)
    best_ws = None
    best_table = None
    best_count = -1
    for ws in wb.worksheets:
        table = detect_table(ws, "order")
        if not table or "order_qty" not in table.columns:
            continue
        count = 0
        for r in range(table.data_start, last_nonempty_row(ws) + 1):
            if to_number(ws.cell(r, table.columns["order_qty"]).value):
                count += 1
        if count > best_count:
            best_ws, best_table, best_count = ws, table, count
    return wb, best_ws, best_table


def _best_shipment_sheet(path: Path, store: str, products: dict[str, dict[str, Any]]):
    wb = load_workbook(path)
    desired_keys = set()
    for item in products.values():
        desired_keys.update(_shipment_match_keys(item))
    store_key = normalize_key(store)
    best_ws = None
    best_table = None
    best_score = -1
    for index, ws in enumerate(wb.worksheets):
        table = detect_table(ws, "order")
        if not table or "order_qty" not in table.columns:
            continue
        score = 0
        title_key = normalize_key(ws.title)
        if store_key and store_key in title_key:
            score += 1000
        if "code" in table.columns and "warehouse" in table.columns:
            score += 50
        for row in range(table.data_start, last_nonempty_row(ws) + 1):
            row_keys = _shipment_row_match_keys(ws, row, table.columns)
            if any(key in desired_keys for key in row_keys):
                score += 10
        score -= index
        if score > best_score:
            best_ws, best_table, best_score = ws, table, score
    return wb, best_ws, best_table


def update_store_header(ws, store: str) -> None:
    fallback_title = f"{store}发货单"
    for row in range(1, min(6, ws.max_row) + 1):
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row, col).value
            if isinstance(value, str) and any(token in value for token in ("馄饨侯", "发货", "出库", "店")):
                title = value.replace("产品订货单", "产品发货单").replace("订货单", "发货单").replace("出库单", "发货单")
                ws.cell(row, col).value = title if title != value or "发货单" in title else fallback_title
                return
    table = detect_table(ws, "order")
    if table and table.header_row <= 2:
        return
    ws["A1"].value = fallback_title


def _store_title_name(store: str) -> str:
    text = str(store or "").strip() or "门店"
    if text.endswith("店"):
        return text[:-1]
    return text


def _safe_sheet_title(value: Any) -> str:
    text = re.sub(r"[\[\]:*?/\\]", "_", str(value or "门店")).strip()
    text = text.strip("'") or "门店"
    return text[:31]


def update_order_header(ws, store: str, order_date: date | None, items: list[dict[str, Any]]) -> None:
    store_name = _store_title_name(store)
    fallback_title = f"馄饨侯（{store_name}）店产品订货单"
    for row in range(1, min(6, ws.max_row) + 1):
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row, col).value
            if not isinstance(value, str):
                continue
            if any(token in value for token in ("馄饨侯", "订货单", "发货单")) and "产品" in value:
                ws.cell(row, col).value = fallback_title
                break

    order_date_text = order_date.isoformat() if order_date else ""
    delivery_date = first_present(*(item.get("deliver_date") for item in items))
    orderer = first_present(
        *(item.get(field) for item in items for field in ("orderer", "buyer", "contact_name", "customer"))
    )
    phone = first_present(*(item.get(field) for item in items for field in ("phone", "contact_phone", "mobile")))

    label_values = {
        "订货日期": order_date_text,
        "到货日期": delivery_date,
        "订货人": orderer,
        "联系电话": phone,
    }
    for row in range(1, min(6, ws.max_row) + 1):
        row_values = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
        for label, target_value in label_values.items():
            label_col = None
            for col, value in enumerate(row_values, 1):
                if isinstance(value, str) and label in value:
                    label_col = col
                    break
            if not label_col:
                continue
            target_col = None
            for col in range(label_col + 1, ws.max_column + 1):
                value = ws.cell(row, col).value
                if value not in (None, "") and not (isinstance(value, str) and any(name in value for name in label_values)):
                    target_col = col
                    break
            if target_col is None:
                target_col = min(label_col + 1, ws.max_column)
            ws.cell(row, target_col).value = target_value


def _shipment_key(product: Any) -> str:
    return normalize_key(product)


def _shipment_match_keys(item: dict[str, Any]) -> list[str]:
    keys: list[str] = []
    code_key = normalize_key(item.get("code"))
    if code_key:
        keys.append(f"code:{code_key}")
    product_key = normalize_key(item.get("product") or item.get("name"))
    if product_key:
        keys.append(f"product:{product_key}")
    return keys


def _shipment_row_match_keys(ws, row: int, columns: dict[str, int]) -> list[str]:
    keys: list[str] = []
    code_col = columns.get("code")
    if code_col:
        code_key = normalize_key(ws.cell(row, code_col).value)
        if code_key:
            keys.append(f"code:{code_key}")
    product_col = columns.get("product")
    if product_col:
        product_key = normalize_key(ws.cell(row, product_col).value)
        if product_key:
            keys.append(f"product:{product_key}")
    return keys


def _shipment_item_lookup(products: dict[str, dict[str, Any]]) -> dict[str, dict[str, Any]]:
    lookup: dict[str, dict[str, Any]] = {}
    for item in products.values():
        for key in _shipment_match_keys(item):
            lookup.setdefault(key, item)
    return lookup


def _merge_shipment_item(
    store_items: dict[str, dict[str, Any]],
    item: dict[str, Any],
    quantity_value: Any,
) -> None:
    product = str(item.get("product") or item.get("name") or "").strip()
    qty = to_number(quantity_value)
    if not product or qty is None or qty <= 0:
        return
    key = _shipment_key(product)
    current = store_items.setdefault(
        key,
        {
            "product": product,
            "category": "",
            "code": "",
            "spec": "",
            "unit": "",
            "price": None,
            "quantity": 0.0,
        },
    )
    current["quantity"] += float(qty)
    for field in ("category", "code", "spec", "unit"):
        if not current.get(field) and item.get(field) not in (None, ""):
            current[field] = str(item.get(field)).strip()
    for field in ("deliver_date", "orderer", "buyer", "contact_name", "customer", "phone", "contact_phone", "mobile"):
        if not current.get(field) and item.get(field) not in (None, ""):
            current[field] = str(item.get(field)).strip()
    price = to_number(item.get("price"))
    if current.get("price") is None and price is not None:
        current["price"] = price


def _write_shipment_item_to_row(ws, row: int, columns: dict[str, int], item: dict[str, Any], sequence: int | None = None) -> None:
    values = {
        "category": item.get("category", ""),
        "code": item.get("code", ""),
        "product": item.get("product", ""),
        "spec": item.get("spec", ""),
        "unit": item.get("unit", ""),
        "price": display_number(item.get("price")),
        "order_qty": display_number(item.get("quantity")),
    }
    if sequence is not None:
        values["sequence"] = sequence
    for key, value in values.items():
        col = columns.get(key)
        if col:
            ws.cell(row, col).value = value


def _write_shipment_quantity_to_row(ws, row: int, columns: dict[str, int], item: dict[str, Any]) -> None:
    qty_col = columns.get("order_qty")
    if qty_col:
        ws.cell(row, qty_col).value = display_number(item.get("quantity"))


def _write_shipment_matched_row(ws, row: int, columns: dict[str, int], item: dict[str, Any]) -> None:
    values = {
        "category": item.get("category", ""),
        "code": item.get("code", ""),
        "product": item.get("product", ""),
        "spec": item.get("spec", ""),
        "unit": item.get("unit", ""),
        "price": display_number(item.get("price")),
    }
    for key, value in values.items():
        col = columns.get(key)
        if not col or value in (None, ""):
            continue
        if ws.cell(row, col).value in (None, ""):
            ws.cell(row, col).value = value
    _write_shipment_quantity_to_row(ws, row, columns, item)


def _append_shipment_item(ws, table: TableMap, item: dict[str, Any], sequence: int) -> None:
    dst_row = last_nonempty_row(ws) + 1
    src_row = max(table.data_start, dst_row - 1)
    copy_row_format(ws, src_row, dst_row, ws.max_column)
    if ws.row_dimensions[src_row].height:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    _write_shipment_item_to_row(ws, dst_row, table.columns, item, sequence)


def _is_shipment_product_row(ws, row: int, columns: dict[str, int]) -> bool:
    product_col = columns.get("product")
    return bool(product_col and normalize_text(ws.cell(row, product_col).value))


def _renumber_shipment_rows(ws, table: TableMap) -> int:
    sequence_col = table.columns.get("sequence")
    if not sequence_col:
        return 0
    sequence = 0
    for row in range(table.data_start, last_nonempty_row(ws) + 1):
        if _is_shipment_product_row(ws, row, table.columns):
            sequence += 1
            ws.cell(row, sequence_col).value = sequence
    return sequence


def _collect_store_products(
    order_paths: list[Path],
    confirmed_items: list[dict[str, Any]] | None,
) -> tuple[dict[str, dict[str, dict[str, Any]]], dict[str, Path]]:
    store_products: dict[str, dict[str, dict[str, Any]]] = defaultdict(dict)
    store_template: dict[str, Path] = {}

    for path in order_paths:
        for row in parse_rows(path, "order"):
            qty = row.get("order_qty")
            if qty is None or qty == 0:
                continue
            _merge_shipment_item(store_products[row["store"]], row, qty)
            store_template.setdefault(row["store"], path)

    for item in confirmed_items or []:
        store = str(item.get("store") or "").strip()
        if not store:
            continue
        _merge_shipment_item(store_products[store], item, item.get("quantity"))

    return store_products, store_template


def generate_shipment_outputs(
    order_paths: list[Path],
    template_path: Path | None,
    confirmed_items: list[dict[str, Any]] | None,
    order_date: date | None,
    output_dir: Path,
) -> tuple[Path, list[str]]:
    output_dir.mkdir(parents=True, exist_ok=True)
    warnings: list[str] = []
    store_products, store_template = _collect_store_products(order_paths, confirmed_items)

    if not store_products:
        warnings.append("没有可生成发货单的订货数量或确认发货文本。")

    generated: list[Path] = []
    output_date = (order_date or date.today()).isoformat()
    first_template = template_path or (order_paths[0] if order_paths else None)
    for store, products in store_products.items():
        template = store_template.get(store) or first_template
        if not template:
            wb = Workbook()
            ws = wb.active
            write_basic_headers(ws, f"{store}发货单", ["序号", "类别", "编码", "原料名称", "规格", "单位", "单价", "订货数量"])
            for idx, item in enumerate(products.values(), 1):
                ws.append(
                    [
                        idx,
                        item.get("category", ""),
                        item.get("code", ""),
                        item.get("product", ""),
                        item.get("spec", ""),
                        item.get("unit", ""),
                        display_number(item.get("price")),
                        display_number(item.get("quantity")),
                    ]
                )
        else:
            wb, ws, table = _best_shipment_sheet(template, store, products)
            if not ws or not table:
                warnings.append(f"{template.name} 没有识别到订单格式，已跳过。")
                continue
            for other in list(wb.worksheets):
                if other is not ws:
                    wb.remove(other)
            ws.title = _safe_sheet_title(store)
            update_store_header(ws, store)
            qty_col = table.columns.get("order_qty")
            lookup = _shipment_item_lookup(products)
            seen_items: set[int] = set()
            rows_to_delete: list[int] = []
            if qty_col:
                for r in range(table.data_start, last_nonempty_row(ws) + 1):
                    item = None
                    for key in _shipment_row_match_keys(ws, r, table.columns):
                        item = lookup.get(key)
                        if item:
                            break
                    if item:
                        seen_items.add(id(item))
                        _write_shipment_matched_row(ws, r, table.columns, item)
                    elif _is_shipment_product_row(ws, r, table.columns):
                        rows_to_delete.append(r)
                    else:
                        ws.cell(r, qty_col).value = None
            for row in reversed(rows_to_delete):
                ws.delete_rows(row)
            max_sequence = _renumber_shipment_rows(ws, table)
            for item in products.values():
                if id(item) in seen_items:
                    continue
                max_sequence += 1
                _append_shipment_item(ws, table, item, max_sequence)
            _renumber_shipment_rows(ws, table)
        safe_store = re.sub(r"[^\w\u4e00-\u9fff]+", "_", store).strip("_") or "门店"
        output = output_dir / f"{safe_store}_发货单_{output_date}.xlsx"
        wb.save(output)
        generated.append(output)

    if len(generated) == 1:
        return generated[0], warnings
    zip_path = output_dir / f"发货单_{output_date}.zip"
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for path in generated:
            archive.write(path, path.name)
    return zip_path, warnings


def generate_order_documents(
    order_paths: list[Path],
    template_path: Path | None,
    confirmed_items: list[dict[str, Any]] | None,
    order_date: date | None,
    output_dir: Path,
) -> tuple[Path, list[str]]:
    output_dir.mkdir(parents=True, exist_ok=True)
    warnings: list[str] = []
    store_products, store_template = _collect_store_products(order_paths, confirmed_items)

    if not store_products:
        warnings.append("没有可生成订货单的订货数量或确认文本。")

    generated: list[Path] = []
    output_date = (order_date or date.today()).isoformat()
    for store, products in store_products.items():
        wb = _build_order_document_workbook(store, list(products.values()), order_date)
        safe_store = re.sub(r"[^\w\u4e00-\u9fff]+", "_", store).strip("_") or "门店"
        output = output_dir / f"{safe_store}_订货单_{output_date}.xlsx"
        wb.save(output)
        generated.append(output)

    if len(generated) == 1:
        return generated[0], warnings
    zip_path = output_dir / f"订货单_{output_date}.zip"
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for path in generated:
            archive.write(path, path.name)
    return zip_path, warnings


def _format_order_date(value: date | None) -> str:
    if not value:
        return ""
    return f"{value.month}/{value.day}/{value.year}"


def _build_order_document_workbook(store: str, items: list[dict[str, Any]], order_date: date | None) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_title(store)

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    header_fill = PatternFill("solid", fgColor="F3F3F3")

    ws.merge_cells("A1:H1")
    ws["A1"] = f"馄饨侯（{_store_title_name(store)}）店产品订货单"
    ws["A1"].font = Font(name="SimSun", size=18, bold=True)
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 32

    orderer = first_present(*(item.get(field) for item in items for field in ("orderer", "buyer", "contact_name", "customer")))
    phone = first_present(*(item.get(field) for item in items for field in ("phone", "contact_phone", "mobile")))
    delivery_date = first_present(*(item.get("deliver_date") for item in items))

    for cell, value in {
        "A2": "订货日期：",
        "D2": _format_order_date(order_date),
        "H2": "订货人：",
        "I2": orderer,
        "A3": "到货日期：",
        "D3": delivery_date,
        "H3": "联系电话：",
        "I3": phone,
    }.items():
        ws[cell] = value
        ws[cell].font = Font(name="SimSun", size=12, bold=cell in {"A2", "A3", "H2", "H3"})
        ws[cell].alignment = left if cell in {"A2", "A3", "H2", "H3"} else center

    headers = ["序号", "类别", "编码", "原料名称", "规格", "单位", "单价", "订货数量"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(4, col)
        cell.value = header
        cell.font = Font(name="SimSun", size=11, bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    for row_index, item in enumerate(items, 5):
        values = [
            row_index - 4,
            item.get("category", ""),
            item.get("code", ""),
            item.get("product", ""),
            item.get("spec", ""),
            item.get("unit", ""),
            display_number(item.get("price")),
            display_number(item.get("quantity")),
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row_index, col)
            cell.value = value
            cell.font = Font(name="SimSun", size=11)
            cell.border = border
            cell.alignment = center if col in {1, 6, 7, 8} else left

    widths = {1: 8, 2: 12, 3: 14, 4: 24, 5: 20, 6: 10, 7: 12, 8: 12, 9: 16}
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A5"
    return wb


def _eval_linear_formula(ws, cell_ref: str, cache: dict[str, tuple[float, float]], stack: set[str] | None = None) -> tuple[float, float]:
    stack = stack or set()
    cell_ref = cell_ref.replace("$", "").upper()
    if cell_ref in cache:
        return cache[cell_ref]
    if cell_ref in stack:
        return (1.0, 0.0)
    stack.add(cell_ref)
    value = ws[cell_ref].value
    if value in (None, ""):
        stack.remove(cell_ref)
        return (1.0, 0.0)
    if isinstance(value, (int, float)):
        stack.remove(cell_ref)
        return (0.0, float(value))
    expr = str(value).strip()
    if expr.startswith("="):
        expr = expr[1:]
    expr = expr.replace("$", "").upper()
    expr = expr.replace("L3", "Q")

    def repl(match: re.Match[str]) -> str:
        ref = f"{match.group(1)}{match.group(2)}"
        if ref == "L3":
            return "Q"
        a, b = _eval_linear_formula(ws, ref, cache, stack)
        return f"(({a})*Q+({b}))"

    expr = CELL_REF_RE.sub(repl, expr)
    if not re.fullmatch(r"[0-9Qq+\-*/(). ]+", expr):
        stack.remove(cell_ref)
        return (1.0, 0.0)

    def evaluate(qty: float) -> float:
        return float(eval(expr.replace("Q", str(qty)), {"__builtins__": {}}, {}))

    try:
        b = evaluate(0)
        a = evaluate(1) - b
    except Exception:
        a, b = 1.0, 0.0
    cache[cell_ref] = (a, b)
    stack.remove(cell_ref)
    return a, b


def _parse_feed_sheet(ws) -> list[dict[str, Any]]:
    title_name = clean_finished_name(ws.title)
    first_cell_name = clean_finished_name(ws["A1"].value)
    finished = title_name or first_cell_name
    if not finished:
        return []
    rows: list[dict[str, Any]] = []
    formula_cache: dict[str, tuple[float, float]] = {}
    last = last_nonempty_row(ws)
    header_rows = []
    for r in range(1, last + 1):
        headers = [normalize_text(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if "原料名称" in headers and any("单品净重" in value for value in headers) and "得率" in headers:
            header_rows.append(r)

    for header_row in header_rows:
        for r in range(header_row + 1, last + 1):
            marker = normalize_text(ws.cell(r, 1).value)
            raw_name = normalize_text(ws.cell(r, 2).value)
            if not raw_name:
                continue
            if raw_name == "原料名称":
                break
            if marker.upper() == "TTL":
                break
            net_weight_g = to_number(ws.cell(r, 3).value)
            yield_rate = to_number(ws.cell(r, 4).value) or 1.0
            if net_weight_g is None or not yield_rate:
                continue
            qty_per_unit_kg = float(net_weight_g) / float(yield_rate) / 1000
            units_multiplier, units_addend = _eval_linear_formula(ws, f"G{r}", formula_cache)
            rows.append(
                {
                    "finished_key": normalize_key(finished),
                    "finished": finished,
                    "raw_key": normalize_key(raw_name),
                    "raw": raw_name,
                    "unit": "kg",
                    "spec": "",
                    "code": "",
                    "qty": qty_per_unit_kg,
                    "units_multiplier": units_multiplier,
                    "units_addend": units_addend,
                    "source_sheet": ws.title,
                    "source_row": r,
                }
            )
    return rows


def parse_recipe_table(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=False)
    feed_rows: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        feed_rows.extend(_parse_feed_sheet(ws))
    if feed_rows:
        return feed_rows

    rows = parse_rows(path, "generic")
    recipes = []
    for row in rows:
        finished = row.get("category") or row.get("product")
        raw_name = row.get("product")
        qty = row.get("qty") or row.get("order_qty")
        if not finished or not raw_name or qty is None:
            continue
        recipes.append(
            {
                "finished_key": normalize_key(finished),
                "finished": finished,
                "raw_key": normalize_key(raw_name),
                "raw": raw_name,
                "unit": row.get("unit", ""),
                "spec": row.get("spec", ""),
                "code": row.get("code", ""),
                "qty": float(qty),
            }
        )
    return recipes


def parse_recipe_tables(paths: list[Path]) -> list[dict[str, Any]]:
    recipes: list[dict[str, Any]] = []
    for path in paths:
        recipes.extend(parse_recipe_table(path))
    return recipes


def summarize_recipe_tables(paths: list[Path]) -> dict[str, Any]:
    files = []
    total_rows = 0
    total_products: Counter[str] = Counter()
    for path in paths:
        workbook_rows = 0
        workbook_products: Counter[str] = Counter()
        unrecognized_sheets: list[str] = []
        try:
            wb = load_workbook(path, data_only=False)
        except Exception as exc:
            files.append(
                {
                    "name": path.name,
                    "recipe_rows": 0,
                    "products": [],
                    "unrecognized_sheets": [],
                    "error": str(exc),
                }
            )
            continue

        for ws in wb.worksheets:
            rows = _parse_feed_sheet(ws)
            if rows:
                workbook_rows += len(rows)
                workbook_products.update(row["finished"] for row in rows)
            elif last_nonempty_row(ws) > 1 or normalize_text(ws["A1"].value):
                unrecognized_sheets.append(ws.title)

        if workbook_rows == 0:
            fallback_rows = parse_recipe_table(path)
            workbook_rows = len(fallback_rows)
            workbook_products.update(row["finished"] for row in fallback_rows)

        total_rows += workbook_rows
        total_products.update(workbook_products)
        files.append(
            {
                "name": path.name,
                "recipe_rows": workbook_rows,
                "products": [{"name": name, "rows": count} for name, count in workbook_products.most_common()],
                "unrecognized_sheets": unrecognized_sheets,
            }
        )

    return {
        "file_count": len(paths),
        "recipe_rows": total_rows,
        "product_count": len(total_products),
        "products": [{"name": name, "rows": count} for name, count in total_products.most_common()],
        "files": files,
    }


def recipe_required_qty(recipe: dict[str, Any], production_qty: float) -> float:
    multiplier = float(recipe.get("units_multiplier", 1.0))
    addend = float(recipe.get("units_addend", 0.0))
    units = multiplier * float(production_qty) + addend
    return float(recipe["qty"]) * max(units, 0.0)


def parse_workshop_stock(text: str) -> dict[str, float]:
    result: dict[str, float] = {}
    for line in re.split(r"[\n,，;；]+", text):
        match = NUMBER_RE.search(line)
        if not match:
            continue
        name = line[: match.start()].strip(" ：:\t")
        if not name:
            name = line[match.end() :].strip(" ：:\t")
        if name:
            result[normalize_key(name)] = float(match.group())
    return result


def parse_conversion_table(path: Path | None) -> dict[str, float]:
    if not path:
        return {}
    conversions: dict[str, float] = {}
    for row in parse_rows(path, "generic"):
        qty = row.get("qty") or row.get("order_qty")
        if qty:
            conversions[row["product_key"]] = float(qty)
    return conversions


_SPEC_UNIT_TO_KG = {
    "kg": 1.0,
    "KG": 1.0,
    "千克": 1.0,
    "公斤": 1.0,
    "斤": 0.5,
    "g": 0.001,
    "G": 0.001,
    "克": 0.001,
    "l": 1.0,
    "L": 1.0,
    "升": 1.0,
    "ml": 0.001,
    "ML": 0.001,
    "毫升": 0.001,
}


def _spec_conversion_factor(detail: dict[str, Any]) -> float | None:
    unit = normalize_text(detail.get("unit"))
    if unit in _SPEC_UNIT_TO_KG:
        return _SPEC_UNIT_TO_KG[unit]

    spec = normalize_text(detail.get("spec"))
    if spec:
        match = re.search(r"(\d+(?:\.\d+)?)(kg|KG|千克|公斤|斤|g|G|克|ml|ML|毫升|l|L|升)", spec)
        if match:
            unit = match.group(2)
            return float(match.group(1)) * _SPEC_UNIT_TO_KG[unit]
    return None


def _coerce_paths(value: Path | list[Path] | None) -> list[Path]:
    if not value:
        return []
    if isinstance(value, list):
        return [path for path in value if path]
    return [value]


def _clean_owner_value(value: Any) -> str:
    text = normalize_text(value)
    if text in {"(空白)", "空白", "（空白）"}:
        return ""
    return text


def _owner_detail_score(detail: dict[str, Any]) -> int:
    score = sum(1 for field in ("warehouse", "code", "spec", "unit") if detail.get(field))
    if detail.get("price") is not None:
        score += 1
    return score


def parse_stock_owner_details(path: Path | list[Path] | None) -> dict[str, dict[str, Any]]:
    paths = _coerce_paths(path)
    if not paths:
        return {}
    owners: dict[str, dict[str, Any]] = {}
    for path_item in paths:
        for row in parse_rows(path_item, "generic"):
            key = row.get("product_key")
            if not key:
                continue
            detail = {
                "warehouse": _clean_owner_value(row.get("warehouse")),
                "code": _clean_owner_value(row.get("code")),
                "spec": _clean_owner_value(row.get("spec")),
                "unit": _clean_owner_value(row.get("unit")),
                "price": row.get("price"),
            }
            existing = owners.get(key)
            if not existing or _owner_detail_score(detail) >= _owner_detail_score(existing):
                owners[key] = detail
    return owners


def parse_stock_owner_table(path: Path | list[Path] | None) -> dict[str, str]:
    return {
        key: detail["warehouse"]
        for key, detail in parse_stock_owner_details(path).items()
        if detail.get("warehouse")
    }


def _recipe_candidates_for_product(
    recipe_by_product: dict[str, list[dict[str, Any]]],
    product_key: str,
    product_name: str,
) -> tuple[list[dict[str, Any]], str | None]:
    exact = recipe_by_product.get(product_key, [])
    if exact:
        return exact, None

    matched_keys = [
        finished_key
        for finished_key in recipe_by_product
        if product_key and finished_key and (finished_key in product_key or product_key in finished_key)
    ]
    if not matched_keys:
        return [], None

    matched_names = []
    for key in matched_keys:
        first = recipe_by_product[key][0]
        matched_names.append(str(first.get("finished") or key))
    if len(matched_keys) > 1:
        names = "、".join(matched_names)
        return [], f"{product_name} 模糊匹配到多个投料单：{names}。请确认后调整商品名或配方表。"
    key = matched_keys[0]
    matched_name = matched_names[0]
    return recipe_by_product[key], f"{product_name} 未精确命中投料单，已按模糊匹配使用 {matched_name}。"


def generate_material_issue_workbook(
    production_path: Path | None,
    recipe_paths: list[Path],
    conversion_path: Path | None,
    stock_owner_path: Path | list[Path] | None,
    material_template_path: Path | None,
    workshop_stock_text: str,
    document_date: date | None,
    output_dir: Path,
) -> tuple[Path | None, list[str], list[str]]:
    missing = []
    if not production_path:
        missing.append("已填好的排产表")
    if not recipe_paths:
        missing.append("原材料配方表/投料单")
    if not stock_owner_path:
        missing.append("所属库表")
    if missing:
        return None, missing, []

    warnings: list[str] = []
    production_rows = parse_rows(production_path, "production")
    recipes = parse_recipe_tables(recipe_paths)
    stock_owner_details = parse_stock_owner_details(stock_owner_path)
    workshop_stock = parse_workshop_stock(workshop_stock_text)

    recipe_by_product: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for recipe in recipes:
        recipe_by_product[recipe["finished_key"]].append(recipe)

    material_qty: dict[str, dict[str, Any]] = {}
    for row in production_rows:
        safety = row.get("safety")
        if safety is None:
            continue
        inventory = row.get("inventory") or 0.0
        inbound = row.get("inbound") or 0.0
        outbound = row.get("outbound") or 0.0
        current = row.get("theory_stock")
        if current is None:
            current = float(inventory) + float(inbound) - float(outbound)
        current = float(current)
        if current >= float(safety) * 0.5:
            continue
        production_qty = float(row.get("production") or safety)
        product_recipes, fuzzy_warning = _recipe_candidates_for_product(recipe_by_product, row["product_key"], row["product"])
        if fuzzy_warning:
            warnings.append(fuzzy_warning)
        if not product_recipes:
            if not fuzzy_warning:
                warnings.append(f"{row['product']} 触发领料，但没有找到对应投料单/配方。")
            continue
        for recipe in product_recipes:
            raw_key = recipe["raw_key"]
            need = recipe_required_qty(recipe, production_qty)
            need -= workshop_stock.get(raw_key, 0.0)
            if need <= 0:
                continue
            owner = stock_owner_details.get(raw_key, {})
            factor = _spec_conversion_factor(owner)
            issue_qty = math.ceil(need / factor) if factor else need
            if not factor:
                warnings.append(f"{recipe['raw']} 在所属库表中没有可换算规格，按原始用量输出。")
            owner_price = owner.get("price")
            current_item = material_qty.setdefault(
                raw_key,
                {
                    "code": first_present(owner.get("code"), recipe.get("code")),
                    "raw": recipe["raw"],
                    "spec": first_present(owner.get("spec"), recipe.get("spec")),
                    "unit": first_present(owner.get("unit"), recipe.get("unit")),
                    "warehouse": owner.get("warehouse", ""),
                    "price": owner_price if owner_price is not None else to_number(recipe.get("price")),
                    "qty": 0.0,
                },
            )
            for field in ("code", "spec", "unit", "warehouse"):
                if not current_item.get(field) and owner.get(field):
                    current_item[field] = owner[field]
            if current_item.get("price") is None and owner_price is not None:
                current_item["price"] = owner_price
            if not current_item["warehouse"]:
                warnings.append(f"{recipe['raw']} 没有在所属库表中找到所属库。")
            current_item["qty"] += issue_qty

    output_dir.mkdir(parents=True, exist_ok=True)
    today = (document_date or date.today()).isoformat()
    if material_template_path:
        wb = load_workbook(material_template_path)
        ws = wb.worksheets[0]
        table = detect_table(ws, "material")
        data_start = table.data_start if table else 2
        cols = table.columns if table else {"code": 1, "product": 2, "spec": 3, "unit": 4, "qty": 5}
        if "warehouse" not in cols:
            cols["warehouse"] = max(ws.max_column, max(cols.values(), default=0)) + 1
            ws.cell((table.header_row if table else 1), cols["warehouse"]).value = "所属库"
        if "price" not in cols:
            cols["price"] = max(ws.max_column, max(cols.values(), default=0)) + 1
            ws.cell((table.header_row if table else 1), cols["price"]).value = "单价"
        max_col = max(ws.max_column, cols["warehouse"], cols["price"], 14)
        for r in range(data_start, max(last_nonempty_row(ws), data_start + len(material_qty) + 5) + 1):
            copy_row_format(ws, data_start, r, max_col)
            for c in range(1, max_col + 1):
                ws.cell(r, c).value = None
        for offset, item in enumerate(material_qty.values()):
            r = data_start + offset
            for key, value in {
                "code": item["code"],
                "product": item["raw"],
                "spec": item["spec"],
                "unit": item["unit"],
                "qty": display_number(item["qty"]),
                "warehouse": item["warehouse"],
                "price": display_number(item["price"]),
            }.items():
                col = cols.get(key)
                if col:
                    ws.cell(r, col).value = value
    else:
        wb = Workbook()
        ws = wb.active
        write_basic_headers(ws, "材料出库单", ["存货编码", "存货名称", "规格型号", "主计量", "数量", "所属库", "单价"])
        for item in material_qty.values():
            ws.append([item["code"], item["raw"], item["spec"], item["unit"], display_number(item["qty"]), item["warehouse"], display_number(item["price"])])

    output = output_dir / f"材料出库单_{today}.xlsx"
    wb.save(output)
    return output, missing, warnings


def generate_receipt_workbook(
    receipt_template_path: Path | None,
    items: list[dict[str, Any]],
    document_date: date | None,
    output_dir: Path,
) -> tuple[Path, list[str]]:
    output_dir.mkdir(parents=True, exist_ok=True)
    warnings: list[str] = []
    today = (document_date or date.today()).isoformat()
    clean_items = []
    for item in items:
        name = str(item.get("product") or item.get("name") or "").strip()
        qty = to_number(item.get("quantity"))
        if not name or qty is None:
            continue
        clean_items.append(
            {
                "code": item.get("code", ""),
                "product": name,
                "spec": item.get("spec", ""),
                "unit": item.get("unit", ""),
                "quantity": qty,
            }
        )
    if not clean_items:
        warnings.append("没有可写入入库单的产成品行。")
    if receipt_template_path:
        wb = load_workbook(receipt_template_path)
        ws = wb.worksheets[0]
        table = detect_table(ws, "receipt")
        data_start = table.data_start if table else 2
        cols = table.columns if table else {"code": 1, "product": 2, "spec": 3, "unit": 4, "qty": 5}
        max_col = max(ws.max_column, 12)
        for r in range(data_start, max(last_nonempty_row(ws), data_start + len(clean_items) + 5) + 1):
            copy_row_format(ws, data_start, r, max_col)
            for c in range(1, max_col + 1):
                ws.cell(r, c).value = None
        for offset, item in enumerate(clean_items):
            r = data_start + offset
            for key, value in {
                "code": item["code"],
                "product": item["product"],
                "spec": item["spec"],
                "unit": item["unit"],
                "qty": display_number(item["quantity"]),
            }.items():
                col = cols.get(key)
                if col:
                    ws.cell(r, col).value = value
    else:
        wb = Workbook()
        ws = wb.active
        write_basic_headers(ws, "产成品入库单", ["存货编码", "存货名称", "规格型号", "主计量", "数量"])
        for item in clean_items:
            ws.append([item["code"], item["product"], item["spec"], item["unit"], display_number(item["quantity"])])
    output = output_dir / f"产成品入库单_{today}.xlsx"
    wb.save(output)
    return output, warnings
