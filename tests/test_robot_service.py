import asyncio
from datetime import date
from tempfile import TemporaryDirectory
from pathlib import Path

from openpyxl import Workbook, load_workbook

from backend.services import robot_service
from backend.services.excel_service import (
    generate_completed_production_workbook,
    generate_material_issue_workbook,
    generate_order_documents,
    generate_production_workbook,
    generate_shipment_outputs,
    last_nonempty_row,
)
from backend.services import robot_marking
from backend.services.robot_service import normalize_robot_orders, normalize_robot_receipts


def test_normalize_robot_orders_groups_by_order_date_and_rejects_patch_without_base() -> None:
    payload = {
        "orders": [
            {
                "id": 123,
                "kind": "base",
                "source": "excel",
                "store": "鼓楼店",
                "order_no": "A001",
                "order_date": "2026-06-21",
                "deliver_date": "2026-06-22",
                "items": [
                    {
                        "code": "05020094",
                        "name": "鸡汤虾肉馄饨",
                        "spec": "500g/袋*12袋",
                        "unit": "箱",
                        "qty": 1,
                        "price": 399.11,
                        "category": "馄饨",
                    }
                ],
            },
            {
                "id": 456,
                "kind": "patch",
                "source": "text",
                "store": "鼓楼店",
                "change_type": "add",
                "order_date": "2026-06-21",
                "deliver_date": "2026-06-23",
                "items": [{"code": None, "name": "鸡汤虾肉馄饨", "unit": "箱", "qty": 2}],
            },
            {
                "id": 789,
                "kind": "patch",
                "source": "text",
                "store": "老三家",
                "order_date": "2026-06-21",
                "items": [{"code": "#N/A", "name": "鸡腿", "unit": "件", "qty": 20}],
            },
        ]
    }
    result = normalize_robot_orders(payload)
    assert result["ids"] == [123, 456]
    assert result["all_ids"] == [123, 456, 789]
    assert result["order_dates"] == ["2026-06-21"]
    assert "deliver_dates" not in result
    assert "target_deliver_date" not in result
    assert "blocking_reasons" not in result
    assert result["counts"]["orders"] == 3
    assert result["counts"]["items"] == 2
    assert result["counts"]["stores"] == 1
    assert result["counts"]["rejected_patches"] == 1
    assert result["rejected_patches"][0]["store"] == "老三家"
    assert result["rejected_patches"][0]["order_date"] == "2026-06-21"
    assert result["rejected_patches"][0]["items"][0]["label"] == "鸡腿 20件"

    assert len(result["batches"]) == 1
    batch = result["batches"][0]
    assert batch["order_date"] == "2026-06-21"
    assert batch["ids"] == [123, 456]
    assert batch["counts"]["items"] == 2
    gulou = next(group for group in batch["grouped"] if group["store"] == "鼓楼店")
    assert len(gulou["orders"]) == 2
    assert sum(item["quantity"] for item in gulou["items"]) == 3


def test_normalize_robot_orders_accepts_patch_when_uploaded_base_store_exists() -> None:
    payload = {
        "orders": [
            {
                "id": 789,
                "kind": "patch",
                "source": "text",
                "store": "老三家",
                "order_date": "2026-06-21",
                "items": [{"code": "#N/A", "name": "鸡腿", "unit": "件", "qty": 20}],
            }
        ]
    }
    result = normalize_robot_orders(payload, extra_base_stores={"老三家"})
    assert result["ids"] == [789]
    assert result["rejected_patches"] == []
    assert result["counts"]["items"] == 1
    assert result["batches"][0]["order_date"] == "2026-06-21"
    assert result["batches"][0]["grouped"][0]["store"] == "老三家"


def test_normalize_robot_orders_splits_multiple_order_dates_without_blocking() -> None:
    payload = {
        "orders": [
            {
                "id": 1,
                "kind": "base",
                "store": "A",
                "order_date": "2026-06-21",
                "items": [{"name": "豆浆", "qty": 1, "unit": "箱"}],
            },
            {
                "id": 2,
                "kind": "base",
                "store": "B",
                "order_date": "2026-06-22",
                "items": [{"name": "面条", "qty": 2, "unit": "箱"}],
            },
        ]
    }
    result = normalize_robot_orders(payload)
    assert result["order_dates"] == ["2026-06-21", "2026-06-22"]
    assert [batch["order_date"] for batch in result["batches"]] == ["2026-06-21", "2026-06-22"]
    assert "blocking_reasons" not in result
    assert result["batches"][0]["ids"] == [1]
    assert result["batches"][1]["ids"] == [2]


def test_patch_requires_base_on_same_order_date_when_base_is_from_robot() -> None:
    payload = {
        "orders": [
            {"id": 1, "kind": "base", "store": "鼓楼店", "order_date": "2026-06-21", "items": []},
            {
                "id": 2,
                "kind": "patch",
                "store": "鼓楼店",
                "order_date": "2026-06-22",
                "items": [{"name": "鸡腿", "qty": 20, "unit": "件"}],
            },
        ]
    }
    result = normalize_robot_orders(payload)
    assert result["ids"] == [1]
    assert result["rejected_patches"][0]["id"] == 2
    assert result["rejected_patches"][0]["order_date"] == "2026-06-22"


def test_generate_production_workbook_uses_order_date_for_filename() -> None:
    with TemporaryDirectory() as tmp:
        output, _warnings = generate_production_workbook(
            order_paths=[],
            production_template_path=None,
            safety_stock_path=None,
            confirmed_items=[{"product": "鸡腿", "quantity": 2, "unit": "件"}],
            order_date=date(2026, 6, 21),
            output_dir=Path(tmp),
        )
        assert output.name == "排产表_待补充_2026-06-21.xlsx"
        assert output.exists()
        wb = load_workbook(output, data_only=False)
        ws = wb.active
        assert ws["H3"].value is None
        assert ws["K3"].value is None
        assert ws["L3"].value is None


def test_generate_production_workbook_outputs_only_order_items_from_template() -> None:
    with TemporaryDirectory() as tmp:
        tmp_dir = Path(tmp)
        template_path = tmp_dir / "production_template.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "排产单"
        ws["G2"] = "日期"
        ws.append(["序号", "类别", "编码", "商品名称", "规格", "单位", "单价", "盘点库存数", "安全库存数", "入库数", "出库数量", "理论库存数", "排产量"])
        ws.append([1, "模板", "T1", "模板SKU1", "100g", "箱", 1, None, None, None, None, None, None])
        ws.append([2, "馄饨", "T2", "订单商品", "500g", "箱", 9.5, None, None, None, None, None, None])
        wb.save(template_path)

        output, _warnings = generate_production_workbook(
            order_paths=[],
            production_template_path=template_path,
            safety_stock_path=None,
            confirmed_items=[{"product": "订单商品", "quantity": 3, "unit": "箱"}],
            order_date=date(2026, 6, 21),
            output_dir=tmp_dir,
        )

        wb = load_workbook(output, data_only=False)
        ws = wb.active
        assert ws["C4"].value == "T2"
        assert ws["D4"].value == "订单商品"
        assert ws["K4"].value == 3
        assert ws["D5"].value is None
        products = [ws.cell(row, 4).value for row in range(4, ws.max_row + 1) if ws.cell(row, 4).value]
        assert products == ["订单商品"]


def test_generate_production_workbook_prefers_order_price_when_template_price_is_blank() -> None:
    with TemporaryDirectory() as tmp:
        tmp_dir = Path(tmp)
        template_path = tmp_dir / "production_template.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "排产单"
        ws.append(["序号", "类别", "编码", "商品名称", "规格", "单位", "单价", "盘点库存数", "安全库存数", "入库数", "出库数量", "理论库存数", "排产量"])
        ws.append([1, "烧饼类", "T1", "萝卜丝烧饼", "65g*1", "个", None, None, None, None, None, None, None])
        wb.save(template_path)

        output, _warnings = generate_production_workbook(
            order_paths=[],
            production_template_path=template_path,
            safety_stock_path=None,
            confirmed_items=[{"product": "萝卜丝烧饼", "quantity": 20, "unit": "个", "price": 1.6}],
            order_date=date(2026, 6, 27),
            output_dir=tmp_dir,
        )

        wb = load_workbook(output, data_only=False)
        ws = wb.active
        assert ws["G4"].value == 1.6


def test_generate_production_workbook_fills_safety_from_safety_table() -> None:
    with TemporaryDirectory() as tmp:
        tmp_dir = Path(tmp)
        template_path = tmp_dir / "production_template.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "排产单"
        ws["G2"] = "日期"
        ws.append(["序号", "类别", "编码", "商品名称", "规格", "单位", "单价", "盘点库存数", "安全库存数", "入库数", "出库数量", "理论库存数", "排产量"])
        ws.append([1, "馄饨", "T2", "订单商品", "500g", "箱", 9.5, None, None, None, None, None, None])
        wb.save(template_path)

        safety_path = tmp_dir / "safety.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["随便列", "品名", "库存标准"])
        ws.append(["x", "订单商品", 80])
        wb.save(safety_path)

        output, _warnings = generate_production_workbook(
            order_paths=[],
            production_template_path=template_path,
            safety_stock_path=safety_path,
            confirmed_items=[{"product": "订单商品", "quantity": 3, "unit": "箱"}],
            order_date=date(2026, 6, 21),
            output_dir=tmp_dir,
        )

        wb = load_workbook(output, data_only=False)
        ws = wb.active
        assert ws["I4"].value == 80
        assert ws["K4"].value == 3
        assert ws["L4"].value is None
        assert ws["M4"].value == 80


def test_generate_shipment_uses_order_template_shape_and_full_item_fields() -> None:
    with TemporaryDirectory() as tmp:
        tmp_dir = Path(tmp)
        template_path = tmp_dir / "order_template.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "鼓楼"
        ws["A1"] = "馄饨侯（鼓楼）店产品订货单"
        ws.merge_cells("A1:H1")
        ws["A2"] = "订货日期："
        ws["D2"] = "6/27/2026"
        ws["H2"] = "订货人："
        ws["I2"] = "周凯"
        ws["A3"] = "到货日期："
        ws["H3"] = "联系电话："
        ws["I3"] = "18301369030"
        ws.append(["序号", "类别", "编码", "原料名称", "规格", "单位", "单价", "订货数量"])
        ws.append([1, "馄饨", "05020093", "鸡汤鲜肉馄饨", "260g/袋*25袋", "箱", 267.32, None])
        ws.append([2, "馄饨", "05020094", "鸡汤虾肉馄饨", "500g/袋*12袋", "箱", 399.11, None])
        wb.save(template_path)

        output, warnings = generate_shipment_outputs(
            order_paths=[],
            template_path=template_path,
            confirmed_items=[
                {
                    "store": "鼓楼",
                    "category": "机器人分类不覆盖模板",
                    "code": "BAD-CODE",
                    "product": "鸡汤虾肉馄饨",
                    "spec": "机器人规格不覆盖模板",
                    "unit": "袋",
                    "price": 999.99,
                    "quantity": 3,
                },
                {
                    "store": "鼓楼",
                    "category": "新品",
                    "code": "NEW01",
                    "product": "新增测试品",
                    "spec": "1kg",
                    "unit": "袋",
                    "price": 12.5,
                    "quantity": 2,
                },
            ],
            order_date=date(2026, 6, 27),
            output_dir=tmp_dir,
        )

        assert warnings == []
        wb = load_workbook(output, data_only=True)
        ws = wb.active
        assert ws["A1"].value == "馄饨侯（鼓楼）店产品发货单"
        assert "A1:H1" in [str(item) for item in ws.merged_cells.ranges]
        assert [ws.cell(4, col).value for col in range(1, 9)] == ["序号", "类别", "编码", "原料名称", "规格", "单位", "单价", "订货数量"]
        assert [ws.cell(5, col).value for col in range(1, 9)] == [1, "馄饨", "05020094", "鸡汤虾肉馄饨", "500g/袋*12袋", "箱", 399.11, 3]
        assert [ws.cell(6, col).value for col in range(1, 9)] == [2, "新品", "NEW01", "新增测试品", "1kg", "袋", 12.5, 2]
        assert ws.cell(7, 4).value is None


def test_generate_order_documents_uses_order_template_title_and_filters_rows() -> None:
    with TemporaryDirectory() as tmp:
        tmp_dir = Path(tmp)
        template_path = tmp_dir / "order_template.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "鼓楼"
        ws["A1"] = "馄饨侯（鼓楼）店产品订货单"
        ws.merge_cells("A1:H1")
        ws["A2"] = "订货日期："
        ws["D2"] = "6/27/2026"
        ws["H2"] = "订货人："
        ws["I2"] = "周凯"
        ws["A3"] = "到货日期："
        ws["H3"] = "联系电话："
        ws["I3"] = "18301369030"
        ws.append(["序号", "类别", "编码", "原料名称", "规格", "单位", "单价", "订货数量"])
        ws.append([1, "馄饨", "05020093", "鸡汤鲜肉馄饨", "260g/袋*25袋", "箱", 267.32, None])
        ws.append([2, "馄饨", "05020094", "鸡汤虾肉馄饨", "500g/袋*12袋", "箱", 399.11, None])
        wb.save(template_path)

        output, warnings = generate_order_documents(
            order_paths=[],
            template_path=template_path,
            confirmed_items=[
                {
                    "store": "鼓楼",
                    "product": "鸡汤虾肉馄饨",
                    "spec": "机器人规格不覆盖模板",
                    "unit": "袋",
                    "price": 999.99,
                    "quantity": 3,
                },
                {
                    "store": "鼓楼",
                    "category": "小吃",
                    "code": "NEW01",
                    "product": "新增测试品",
                    "spec": "1kg",
                    "unit": "袋",
                    "price": 12.5,
                    "quantity": 0,
                },
            ],
            order_date=date(2026, 6, 27),
            output_dir=tmp_dir,
        )

        assert warnings == []
        wb = load_workbook(output, data_only=True)
        ws = wb.active
        assert ws["A1"].value == "馄饨侯（鼓楼）店产品订货单"
        assert ws["D2"].value == "2026-06-27"
        assert [ws.cell(5, col).value for col in range(1, 9)] == [1, "馄饨", "05020094", "鸡汤虾肉馄饨", "500g/袋*12袋", "箱", 399.11, 3]
        assert ws.cell(6, 4).value is None


def test_generate_shipment_uses_t6_template_sheet_and_filters_zero_rows() -> None:
    with TemporaryDirectory() as tmp:
        tmp_dir = Path(tmp)
        template_path = tmp_dir / "t6_template.xlsx"
        headers = ["仓库", "存货编码", "存货名称", "规格型号", "主计量", "数量", "含税单价", "税率（%）", "价税合计", "折扣额", "批号", "生产日期", "保质期", "失效日期"]
        wb = Workbook()
        ws = wb.active
        ws.title = "0617鼓楼"
        ws.append(headers)
        ws.append([None, "05020094", "鸡汤虾肉馄饨", "500g/袋*12袋", "箱", 1, None, None, None, None, None, None, None, None])
        ws.append([None, "050200019", "鸡汤菜肉馄饨", "500g/袋*12袋", "箱", 1, None, None, None, None, None, None, None, None])
        ws2 = wb.create_sheet("0617地安门")
        ws2.append(headers)
        ws2.append([None, "05020093", "鸡汤鲜肉馄饨", "280g/袋*25袋", "箱", 4, None, None, None, None, None, None, None, None])
        ws2.append([None, "05020108", "鸡汤鲜虾馄饨", "340g/袋*24袋", "箱", 1, None, None, None, None, None, None, None, None])
        ws2.append([None, "050400059", "鸡棒骨汤", "2kg*1", "袋", 2, None, None, None, None, None, None, None, None])
        wb.save(template_path)

        output, warnings = generate_shipment_outputs(
            order_paths=[],
            template_path=template_path,
            confirmed_items=[
                {
                    "store": "鼓楼",
                    "code": "05020094",
                    "product": "鸡汤虾肉馄饨",
                    "spec": "机器人规格不覆盖模板",
                    "unit": "袋",
                    "price": 399.11,
                    "quantity": 3,
                },
                {
                    "store": "鼓楼",
                    "code": "ZERO",
                    "product": "零数量品",
                    "unit": "箱",
                    "quantity": 0,
                },
                {
                    "store": "鼓楼",
                    "code": "NEW01",
                    "product": "新增品",
                    "spec": "1kg",
                    "unit": "袋",
                    "price": 12.5,
                    "quantity": 2,
                },
            ],
            order_date=date(2026, 6, 27),
            output_dir=tmp_dir,
        )

        assert warnings == []
        wb = load_workbook(output, data_only=True)
        ws = wb.active
        assert ws.title == "鼓楼"
        assert [ws.cell(1, col).value for col in range(1, 15)] == headers
        assert [ws.cell(2, col).value for col in range(1, 8)] == [None, "05020094", "鸡汤虾肉馄饨", "500g/袋*12袋", "箱", 3, 399.11]
        assert [ws.cell(3, col).value for col in range(1, 8)] == [None, "NEW01", "新增品", "1kg", "袋", 2, 12.5]
        exported_names = [ws.cell(row, 3).value for row in range(2, last_nonempty_row(ws) + 1)]
        assert "鸡汤菜肉馄饨" not in exported_names
        assert "鸡汤鲜肉馄饨" not in exported_names
        assert "零数量品" not in exported_names


def test_generate_completed_production_workbook_calculates_theory_stock() -> None:
    with TemporaryDirectory() as tmp:
        tmp_dir = Path(tmp)
        production_path = tmp_dir / "draft.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["序号", "类别", "编码", "商品名称", "规格", "单位", "单价", "盘点库存数", "安全库存数", "入库数", "出库数量", "理论库存数", "理论排产"])
        ws.append([1, "馄饨", "T2", "订单商品", "500g", "箱", 9.5, 20, 80, 5, 3, None, None])
        wb.save(production_path)

        output, warnings = generate_completed_production_workbook(
            production_path=production_path,
            document_date=date(2026, 6, 21),
            output_dir=tmp_dir,
        )

        assert warnings == []
        assert output.name == "排产表_2026-06-21.xlsx"
        wb = load_workbook(output, data_only=False)
        ws = wb.active
        assert ws["L2"].value == 22
        assert ws["M2"].value == 80


def test_generate_material_issue_workbook_adds_warehouse_from_owner_table() -> None:
    with TemporaryDirectory() as tmp:
        tmp_dir = Path(tmp)

        production_path = tmp_dir / "production.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["商品名称", "盘点库存数", "安全库存数", "入库数", "出库数量", "排产量"])
        ws.append(["鸡腿", 10, 100, 0, 0, 100])
        wb.save(production_path)

        recipe_path = tmp_dir / "recipe.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "鸡腿投料单"
        ws.append(["", "", "", "", "", "", ""])
        ws.append(["", "原料名称", "单品净重 g", "得率", "", "", ""])
        ws.append(["", "猪肉馅", 100, 1, "", "", ""])
        wb.save(recipe_path)

        conversion_path = tmp_dir / "conversion.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["存货名称", "数量"])
        ws.append(["猪肉馅", 2])
        wb.save(conversion_path)

        owner_path = tmp_dir / "owner.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["仓库", "存货编码", "存货名称", "规格型号", "计量单位", "本币无税单价"])
        ws.append(["冷冻", "0202", "猪肉馅", "10kg", "斤", 12.25])
        wb.save(owner_path)

        output, missing, warnings = generate_material_issue_workbook(
            production_path=production_path,
            recipe_paths=[recipe_path],
            conversion_path=conversion_path,
            stock_owner_path=owner_path,
            material_template_path=None,
            workshop_stock_text="",
            document_date=date(2026, 6, 21),
            output_dir=tmp_dir,
        )

        assert missing == []
        assert warnings == []
        assert output is not None
        wb = load_workbook(output)
        ws = wb.active
        assert ws["F2"].value == "所属库"
        assert ws["G2"].value == "单价"
        assert ws["A3"].value == "0202"
        assert ws["C3"].value == "10kg"
        assert ws["D3"].value == "斤"
        assert ws["F3"].value == "冷冻"
        assert ws["G3"].value == 12.25


def test_generate_material_issue_workbook_fuzzy_matches_finished_recipe() -> None:
    with TemporaryDirectory() as tmp:
        tmp_dir = Path(tmp)

        production_path = tmp_dir / "production.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["商品名称", "盘点库存数", "安全库存数", "入库数", "出库数量", "排产量"])
        ws.append(["鸡汤虾肉馄饨", 0, 100, 0, 0, 100])
        wb.save(production_path)

        recipe_path = tmp_dir / "recipe.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "虾肉馄饨投料单"
        ws.append(["", "", "", "", "", "", ""])
        ws.append(["", "原料名称", "单品净重 g", "得率", "", "", ""])
        ws.append(["", "虾仁", 100, 1, "", "", ""])
        wb.save(recipe_path)

        conversion_path = tmp_dir / "conversion.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["存货名称", "数量"])
        ws.append(["虾仁", 2])
        wb.save(conversion_path)

        owner_path = tmp_dir / "owner.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["仓库", "存货编码", "存货名称", "规格型号", "计量单位", "本币无税单价"])
        ws.append(["冷冻", "0301", "虾仁", "10kg", "斤", 20])
        wb.save(owner_path)

        output, missing, warnings = generate_material_issue_workbook(
            production_path=production_path,
            recipe_paths=[recipe_path],
            conversion_path=conversion_path,
            stock_owner_path=owner_path,
            material_template_path=None,
            workshop_stock_text="",
            document_date=date(2026, 6, 21),
            output_dir=tmp_dir,
        )

        assert missing == []
        assert any("模糊匹配" in warning and "虾肉馄饨" in warning for warning in warnings)
        assert output is not None
        wb = load_workbook(output)
        ws = wb.active
        assert ws["B3"].value == "虾仁"
        assert ws["F3"].value == "冷冻"


def test_normalize_robot_receipts_summarizes_finished_goods_without_store() -> None:
    result = normalize_robot_receipts(
        {
            "receipts": [
                {
                    "id": "r1",
                    "items": [
                        {"name": "鸡汤虾肉馄饨", "qty": "2", "unit": "箱"},
                        {"name": "鸡汤虾肉馄饨", "qty": 3, "unit": "箱"},
                    ],
                }
            ]
        }
    )
    assert result["ids"] == ["r1"]
    assert result["counts"]["items"] == 2
    assert result["counts"]["products"] == 1
    assert "store" not in result["items"][0]
    assert "grouped" not in result
    assert result["items_summary"][0]["quantity"] == 5


def test_robot_headers_include_bearer_token(monkeypatch) -> None:
    monkeypatch.setattr(robot_service, "ROBOT_API_TOKEN", "shared-token")
    assert robot_service._robot_headers() == {"Authorization": "Bearer shared-token"}


def test_unmark_robot_orders_skips_empty_ids() -> None:
    assert asyncio.run(robot_service.unmark_robot_orders([])) == {"skipped": True, "ids": []}


def test_mark_robot_orders_for_output_clears_successful_shipment_ids(monkeypatch) -> None:
    async def fake_mark(ids):
        return {"succeeded": ids, "failed": [], "ok": True}

    cleared = []
    recorded = []
    monkeypatch.setattr(robot_marking, "mark_robot_orders_fetched", fake_mark)
    monkeypatch.setattr(robot_marking, "clear_robot_mark_failures", lambda ids: cleared.append(ids))
    monkeypatch.setattr(robot_marking, "record_robot_mark_failures", lambda *args: recorded.append(args))

    warnings: list[str] = []
    result = asyncio.run(
        robot_marking.mark_robot_orders_for_output(
            [101, 102],
            warnings,
            {"id": "out-1", "name": "发货单_2026-06-21.xlsx"},
            "发货单",
        )
    )

    assert result == {"succeeded": [101, 102], "failed": [], "ok": True}
    assert warnings == []
    assert cleared == [[101, 102]]
    assert recorded == []


def test_mark_robot_orders_for_output_records_failed_shipment_ids(monkeypatch) -> None:
    async def fake_mark(_ids):
        return {"succeeded": [101], "failed": [102], "ok": False}

    recorded = []
    monkeypatch.setattr(robot_marking, "mark_robot_orders_fetched", fake_mark)
    monkeypatch.setattr(robot_marking, "clear_robot_mark_failures", lambda _ids: None)
    monkeypatch.setattr(robot_marking, "record_robot_mark_failures", lambda *args: recorded.append(args))

    warnings: list[str] = []
    result = asyncio.run(
        robot_marking.mark_robot_orders_for_output(
            [101, 102],
            warnings,
            {"id": "out-1", "name": "发货单_2026-06-21.xlsx"},
            "发货单",
        )
    )

    assert result == {"succeeded": [101], "failed": [102], "ok": False}
    assert warnings == ["发货单已生成，但订单库有 1 个 id 标记失败，可稍后重试：[102]"]
    assert recorded == [
        (
            [102],
            "mark_fetched partial failure",
            {"output_id": "out-1", "output_name": "发货单_2026-06-21.xlsx"},
        )
    ]


def test_generate_shipment_endpoint_is_the_order_mark_call_site() -> None:
    source = Path("backend/main.py").read_text(encoding="utf-8")
    completed_block = source.split('@app.post("/api/generate/production-complete-upload")', 1)[1].split(
        '@app.post("/api/generate/shipment")',
        1,
    )[0]
    shipment_block = source.split('@app.post("/api/generate/shipment")', 1)[1].split(
        '@app.post("/api/generate/material-issue-upload")',
        1,
    )[0]

    assert "mark_robot_orders_for_output" not in completed_block
    assert 'mark_robot_orders_for_output(payload.robot_order_ids, warnings, registered, "发货单")' in shipment_block


def test_mark_robot_receipts_skips_empty_ids() -> None:
    assert asyncio.run(robot_service.mark_robot_receipts_fetched([])) == {"skipped": True, "ids": []}


def test_unmark_robot_receipts_skips_empty_ids() -> None:
    assert asyncio.run(robot_service.unmark_robot_receipts([])) == {"skipped": True, "ids": []}


def test_hard_clear_buttons_and_robot_endpoints_are_wired() -> None:
    index = Path("web/index.html").read_text(encoding="utf-8")
    app_js = Path("web/app.js").read_text(encoding="utf-8")
    main = Path("backend/main.py").read_text(encoding="utf-8")
    robot = Path("backend/services/robot_service.py").read_text(encoding="utf-8")

    assert 'data-hard-clear="orders" data-hard-clear-mode="production"' in index
    assert 'data-hard-clear="receipts" data-hard-clear-mode="receipt"' in index
    assert 'data-hard-clear="orders" data-hard-clear-mode="shipment"' in index
    assert "confirmHardClear" in app_js
    assert '"/api/robot/orders/clear-date"' in app_js
    assert '"/api/robot/receipts/clear-date"' in app_js
    assert '@app.post("/api/robot/orders/clear-date")' in main
    assert '@app.post("/api/robot/receipts/clear-date")' in main
    assert "/api/orders/clear_by_date" in robot
    assert "/api/receipts/clear_by_date" in robot
