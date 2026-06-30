from pathlib import Path

import pytest

from openpyxl import Workbook

from backend.services.excel_service import (
    aggregate_orders,
    extract_receipt_template_skus,
    normalize_key,
    parse_stock_owner_details,
    parse_stock_owner_table,
    parse_recipe_table,
    parse_rows,
    recipe_required_qty,
    safety_stock_map,
    summarize_recipe_tables,
)


def save_order(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "鼓楼"
    ws.append(["馄饨侯（鼓楼）店产品订货单"])
    ws.append([])
    ws.append(["序号", "类别", "编码", "原料名称", "规格", "单位", "单价", "订货数量"])
    ws.append([1, "馄饨", "A1", "鸡汤鲜肉馄饨", "260g", "箱", 10, 2])
    ws.append([2, "馄饨", "A2", "鸡汤虾肉馄饨", "500g", "箱", 12, None])
    wb.save(path)


def save_safety(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["商品名称", "安全库存数"])
    ws.append(["鸡汤鲜肉馄饨", 60])
    wb.save(path)


def test_parse_and_aggregate_order(tmp_path: Path) -> None:
    order = tmp_path / "order.xlsx"
    save_order(order)
    rows = parse_rows(order, "order")
    assert rows[0]["product"] == "鸡汤鲜肉馄饨"
    assert rows[0]["order_qty"] == 2
    summary = aggregate_orders([order])
    item = next(iter(summary.values()))
    assert item["quantity"] == 2


def test_safety_stock_map(tmp_path: Path) -> None:
    safety = tmp_path / "safety.xlsx"
    save_safety(safety)
    values = safety_stock_map(safety)
    assert next(iter(values.values())) == 60


def test_safety_stock_map_does_not_treat_codes_as_stock_counts(tmp_path: Path) -> None:
    safety = tmp_path / "safety_with_mixed_sheets.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "产成品模板"
    ws.append(["产品编码", "产品名称", "规格型号", "计量单位", "安全库存数"])
    ws.append(["050400066", "肉丁炸酱（鼓）", "2kg/袋", "袋", 58])

    bad_sheet = wb.create_sheet("历史单据")
    bad_sheet.append(["商品名称", "商品规格", "单位", "单价（元）", "盘点库存数", "安全库存数"])
    bad_sheet.append(["肉丁炸酱", "2000g*1/袋", "袋", None, None, "050400066"])
    wb.save(safety)

    values = safety_stock_map(safety)

    assert values[normalize_key("050400066")] == 58
    assert values[normalize_key("肉丁炸酱（鼓）")] == 58
    assert values.get(normalize_key("肉丁炸酱")) is None
    assert normalize_key("50400066") not in values


def test_extract_receipt_template_skus_dedupes_and_drops_transaction_fields(tmp_path: Path) -> None:
    path = tmp_path / "receipt_template.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["存货编码", "存货名称", "规格型号", "主计量", "数量", "单价", "金额", "批号", "生产日期"])
    ws.append(["05020122", "鸡蛋面", "", "斤", 10, 2.5, 25, "B1", "2026-06-27"])
    ws.append(["05020122", "鸡蛋面", "", "斤", 30, 2.5, 75, "B2", "2026-06-27"])
    ws.append(["050200014", "麻酱烧饼", "65g*1", "个", 50, 1.6, 80, "B3", "2026-06-27"])
    wb.save(path)

    result = extract_receipt_template_skus(path)

    assert result["source_rows"] == 3
    assert result["unique_rows"] == 2
    assert result["deduped"] == 1
    assert result["products"] == [
        {"name": "鸡蛋面", "spec": "", "unit": "斤", "category": ""},
        {"name": "麻酱烧饼", "spec": "65g*1", "unit": "个", "category": ""},
    ]
    assert all(set(item) == {"name", "spec", "unit", "category"} for item in result["products"])


def test_parse_feed_sheet_recipe(tmp_path: Path) -> None:
    path = tmp_path / "feed.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "鸡蛋面"
    ws["A1"] = "鸡蛋面投料单"
    ws["J3"] = "订单量"
    ws["L3"] = 10
    ws.append([])
    ws.append([])
    ws.append(["NO.", "原料名称", "单品净重  g", "得率", "原料使用量 g", "", "生产个数\n（订单量+保存样）"])
    ws.append([1, "面粉", 100, 0.5, "=C5/D5", "g", "=$L$3+2"])
    ws.append(["TTL", "", "=SUM(C5:C5)"])
    wb.save(path)

    recipes = parse_recipe_table(path)
    assert recipes[0]["finished"] == "鸡蛋面"
    assert recipes[0]["raw"] == "面粉"
    assert recipes[0]["qty"] == 0.2
    assert recipe_required_qty(recipes[0], 10) == pytest.approx(2.4)
    summary = summarize_recipe_tables([path])
    assert summary["file_count"] == 1
    assert summary["product_count"] == 1
    assert summary["recipe_rows"] == 1


def test_parse_stock_owner_table_reads_legacy_xls(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    class FakeSheet:
        name = "所属库"
        rows = [
            ["仓库(cWhName)", "存货编码(cInvCode)", "存货名称(cInvName)", "规格型号(cInvStd)", "计量单位(cInvM_Unit)", "本币无税单价(iUnitCost)"],
            ["主食", "0101", "面粉", "25kg", "袋", 88.5],
            ["冷藏", "0202", "猪肉馅", "10kg", "斤", 12.25],
        ]
        nrows = len(rows)
        ncols = 6

        def cell_value(self, row: int, col: int):
            return self.rows[row][col]

    class FakeBook:
        def sheets(self):
            return [FakeSheet()]

    class FakeXlrd:
        @staticmethod
        def open_workbook(_path: str):
            return FakeBook()

    monkeypatch.setattr("backend.services.excel_service.xlrd", FakeXlrd)
    path = tmp_path / "owners.xls"
    path.write_bytes(b"legacy-xls-placeholder")

    details = parse_stock_owner_details(path)
    owners = parse_stock_owner_table(path)

    assert owners[normalize_key("面粉")] == "主食"
    assert owners[normalize_key("猪肉馅")] == "冷藏"
    assert details[normalize_key("面粉")] == {
        "product": "面粉",
        "product_key": normalize_key("面粉"),
        "warehouse": "主食",
        "code": "0101",
        "spec": "25kg",
        "unit": "袋",
        "price": 88.5,
    }


def test_parse_stock_owner_details_reads_multiple_files_and_prefers_completeness(tmp_path: Path) -> None:
    old_path = tmp_path / "old_owner.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["仓库", "存货名称"])
    ws.append(["主食", "面粉"])
    wb.save(old_path)

    new_path = tmp_path / "new_owner.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["仓库", "存货编码", "存货名称", "规格型号", "计量单位", "本币无税单价"])
    ws.append(["冷冻", "0301", "虾仁", "10kg", "斤", 20])
    ws7 = wb.create_sheet("Sheet7")
    ws7.append(["仓库", "存货编码", "存货名称", "规格型号", "计量单位", "本币无税单价"])
    ws7.append(["调料辅料库", "0101", "面粉", "25kg/袋", "袋", 88.5])
    wb.save(new_path)

    details = parse_stock_owner_details([old_path, new_path])

    assert details[normalize_key("面粉")] == {
        "product": "面粉",
        "product_key": normalize_key("面粉"),
        "warehouse": "调料辅料库",
        "code": "0101",
        "spec": "25kg/袋",
        "unit": "袋",
        "price": 88.5,
    }
    assert details[normalize_key("虾仁")]["code"] == "0301"
    assert details[normalize_key("虾仁")]["price"] == 20
